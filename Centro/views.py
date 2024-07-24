import pandas as pd
import random
import os
from django.shortcuts import get_object_or_404, render, redirect, HttpResponse
from django.contrib import messages
from django.contrib.auth import logout, login
from django.urls import reverse

from django.utils.html import strip_tags

from django.conf import settings
from .models import *
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required

from django.template.loader import get_template, render_to_string
from xhtml2pdf import pisa
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from django.core.mail import send_mail, EmailMultiAlternatives
from Yugoslavia.settings import EMAIL_HOST_USER

# Cargar los códigos postales válidos desde el archivo Excel
excel_path = os.path.join(settings.BASE_DIR, 'Centro', 'data', 'codigo_postal.xlsx')
df = pd.read_excel(excel_path, header=1)
CODIGOS_POSTALES_VALIDOS = df[df['Distrito'] == 'Nuevo Chimbote']['Código Postal'].unique().tolist()
TODOS_CODIGOS_POSTALES = df['Código Postal'].unique().tolist()
def asignar_codigo_postal():
    return random.choice(TODOS_CODIGOS_POSTALES)

def es_codigo_postal_valido(codigo_postal):
    return codigo_postal in CODIGOS_POSTALES_VALIDOS

def obtener_info_codigo_postal(codigo_postal):
    fila = df[df['Código Postal'] == codigo_postal].iloc[0]
    return fila['Provincia'], fila['Departamento']

def login_view(request):
    if request.method == 'POST':
        # Verificar si la solicitud es para el inicio de sesión o el registro
        if 'login_email' in request.POST and 'login_contrasena' in request.POST:
            email = request.POST.get('login_email')
            contrasena = request.POST.get('login_contrasena')
            try:
                useradmin = User.objects.get(email=email)
                if useradmin:
                    messages.success(request, 'Inicio de sesión de administrador exitoso.')
                    login(request, useradmin)
                    request.session['is_logged_in'] = True
                    return redirect('main_admin')
            except User.DoesNotExist:
                pass
            
            paciente = Paciente.objects.filter(Email=email, Contraseña=contrasena).first()

            if paciente:
                return redirect('ver_citas', paciente_id=paciente.id_paciente)
            else:
                messages.error(request, 'Correo electrónico o contraseña incorrectos.')

        else:
            dni = request.POST.get('dni')
            Nombre = request.POST.get('nombre')
            Apellido = request.POST.get('apellido')
            Fecha_nacimiento = request.POST.get('fecha_nacimiento')
            Genero = request.POST.get('genero')
            Direccion = request.POST.get('direccion')
            Telefono = request.POST.get('telefono')
            Email = request.POST.get('email')
            Contraseña = request.POST.get('contrasena')

            # Verificar si el DNI ya está registrado
            if Paciente.objects.filter(Dni=dni).exists():
                messages.error(request, 'El DNI ya está registrado.')
                return redirect('login')

            # Asignar código postal aleatorio
            codigo_postal = asignar_codigo_postal()

            # Verificar si el código postal es válido
            if es_codigo_postal_valido(codigo_postal):
                # Crear un nuevo objeto Paciente
                paciente = Paciente.objects.create(
                    Dni=dni,
                    Nombre=Nombre,
                    Apellido=Apellido,
                    Fecha_nacimiento=Fecha_nacimiento,
                    Genero=Genero,
                    Direccion=Direccion,
                    Telefono=Telefono,
                    Email=Email,
                    Contraseña=Contraseña,
                    Codigo_postal=codigo_postal
                )
                paciente.save()
                messages.success(request, 'Registro exitoso.')
                contexto = {
                    'nombre': paciente.Nombre,
                    'apellido': paciente.Apellido,
                    'dni': paciente.Dni,
                    'email': paciente.Email,
                }
                enviar_email_html(
                    'Registro Exitoso',
                    'registroGmail.html',
                    contexto,
                    paciente.Email
                )
                return redirect('login')
            else:
                provincia, departamento = obtener_info_codigo_postal(codigo_postal)
                return redirect('NoValido', provincia=provincia, departamento=departamento)
    
    return render(request, 'index.html')

def NoValido(request, provincia, departamento):
    context = {
        'provincia': provincia,
        'departamento': departamento
    }
    return render(request, 'NoValido.html', context)

def usuario_citas(request, paciente_id):
    paciente = get_object_or_404(Paciente, id_paciente=paciente_id)
    cita = Cita.objects.filter(paciente=paciente)
    context = {'cita': cita, 'paciente':paciente, 'id':paciente_id}
    return render(request, 'Usuario.html', context)


def usuario_registro(request):
    if request.method == 'POST':
        dni = request.POST.get('dni')
        especialidad = request.POST.get('especialidad')
        medico = request.POST.get('Medico')
        Fechadedisponibilidad = request.POST.get('Fechadedisponibilidad')
        hora = request.POST.get('hora')
        Motivo = request.POST.get('Motivo')
        select_medico = Medico.objects.get(Nombre=medico)
        paciente = Paciente.objects.get(Dni=dni)
        cita = Cita.objects.create(
            medico = select_medico,
            paciente = paciente,
            fecha_cita = Fechadedisponibilidad,
            hora = hora,
            especialidad = especialidad,
            motivo = Motivo
        )
        cita.save()
        contexto = {
            'nombre': paciente.Nombre,
            'especialidad': cita.especialidad,
            'medico': f"{cita.medico.Nombre} {cita.medico.Apellido}",
            'fecha': cita.fecha_cita,
            'hora': cita.hora,
            'motivo': cita.motivo,
        }
        enviar_email_html(
            'Cita Registrada',
            'nuevacitaGmail.html',
            contexto,
            paciente.Email
        )
        messages.success(request, 'Registro exitoso.')
        return redirect(reverse('ver_citas', args=[paciente.id_paciente]))
    Medicos=Medico.objects.all()
    return render(request, 'Rcitas.html', context={'medicos':Medicos})


def cerrar_sesion(request):
    logout(request)
    return redirect('login')

def paciente_pdf(request, paciente_id):
    paciente = get_object_or_404(Paciente, id_paciente=paciente_id)
    citas = Cita.objects.filter(paciente=paciente)
    
    template = get_template('PacientePDF.html')
    context = {
        'paciente': paciente,
        'citas': citas,
    }
    html = template.render(context)
    
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return HttpResponse('Error al generar el PDF', status=400)

@login_required
def main_admin(request):
    citas=Cita.objects.all()
    return render(request, 'Admin.html', context={'citas':citas})

@login_required
def citas_admin(request):
    citas = Cita.objects.all()
    context = {'citas': citas}
    return render(request, 'citas.html', context)

@login_required
def pacientes_admin(request):
    pacientes = Paciente.objects.all()
    print(pacientes)
    context = {'pacientes': pacientes}
    return render(request, 'Pacient.html', context)

@login_required
def medicos_admin(request):
    medicos= Medico.objects.all()
    context = {'medicos':medicos}
    return render(request, 'Especiali.html', context)

def eliminar_cita(request,id_cita):
    cita = Cita.objects.get(id_cita=id_cita)
    cita.delete()
    messages.success(request, 'Cita eliminada.')
    return redirect('citas_admin')

def eliminar_medico(request,id_medico):
    medico = Medico.objects.get(id_medico=id_medico)
    medico.delete()
    messages.success(request, 'Medico eliminado.')
    return redirect('medicos_admin')

def eliminar_paciente(request,id_paciente):
    paciente = Paciente.objects.get(id_paciente=id_paciente)
    paciente.delete()
    messages.success(request, 'Paciente eliminado.')
    return redirect('pacientes_admin')

def crear_paciente(request):
    if request.method == 'POST':
        dni = request.POST.get('dni')
        Nombre = request.POST.get('nombre')
        Apellido = request.POST.get('apellido')
        Fecha_nacimiento = request.POST.get('fecha_nacimiento')
        Genero = request.POST.get('genero')
        Direccion = request.POST.get('direccion')
        Telefono = request.POST.get('telefono')
        Email = request.POST.get('email')
        Contraseña = request.POST.get('contrasena')

        # Verificar si el DNI ya está registrado
        if Paciente.objects.filter(Dni=dni).exists():
            messages.error(request, 'El DNI ya está registrado.')
            return redirect('login')

        # Crear un nuevo objeto Paciente
        paciente = Paciente.objects.create(
            Dni=dni,
            Nombre=Nombre,
            Apellido=Apellido,
            Fecha_nacimiento=Fecha_nacimiento,
            Genero=Genero,
            Direccion=Direccion,
            Telefono=Telefono,
            Email=Email,
            Contraseña=Contraseña
        )
        paciente.save()
        messages.success(request, 'Registro exitoso.')
        return redirect('pacientes_admin')
    return render(request, 'Crear_paciente.html')

def crear_cita(request):
    if request.method == 'POST':
        dni = request.POST.get('dni')
        especialidad = request.POST.get('especialidad')
        medico = request.POST.get('Medico')
        Fechadedisponibilidad = request.POST.get('Fechadedisponibilidad')
        hora = request.POST.get('hora')
        Motivo = request.POST.get('Motivo')
        select_medico = Medico.objects.get(Nombre=medico)
        paciente = Paciente.objects.get(Dni=dni)
        cita = Cita.objects.create(
            medico = select_medico,
            paciente = paciente,
            fecha_cita = Fechadedisponibilidad,
            hora = hora,
            especialidad = especialidad,
            motivo = Motivo
        )
        cita.save()
        messages.success(request, 'Registro exitoso.')
        return redirect('citas_admin')
    Medicos=Medico.objects.all()
    return render(request, 'Crear_cita.html', context={'medicos':Medicos})

def crear_medico(request):
    if request.method == 'POST':
        Nombre = request.POST.get('nombre')
        Apellido = request.POST.get('apellido')
        especialidad = request.POST.get('especialidad')
        Telefono = request.POST.get('telefono')
        Email = request.POST.get('email')
        Contraseña = request.POST.get('contrasena')

        # Crear un nuevo objeto Medico
        medico = Medico.objects.create(
            Nombre=Nombre,
            Apellido=Apellido,
            especialidad=especialidad,
            Telefono=Telefono,
            Email=Email,
            Contraseña=Contraseña
        )
        medico.save()
        messages.success(request, 'Registro exitoso.')
        return redirect('medicos_admin')
    return render(request, 'Crear_medico.html')

def editar_paciente(request,id_paciente):
    paciente = Paciente.objects.get(id_paciente=id_paciente)
    if request.method == 'POST':
        paciente.Dni = request.POST.get('dni')
        paciente.Nombre = request.POST.get('nombre')
        paciente.Apellido = request.POST.get('apellido')
        paciente.Fecha_nacimiento = request.POST.get('fecha_nacimiento')
        paciente.Genero = request.POST.get('genero')
        paciente.Direccion = request.POST.get('direccion')
        paciente.Telefono = request.POST.get('telefono')
        paciente.Email = request.POST.get('email')
        paciente.Contraseña = request.POST.get('contrasena')
        paciente.save()
        messages.success(request, 'Paciente editado.')
        return redirect('pacientes_admin')
    return render(request, 'Editar_paciente.html', context={'paciente':paciente})

def editar_medico(request,id_medico):
    medico = Medico.objects.get(id_medico=id_medico)
    if request.method == 'POST':
        medico.Nombre = request.POST.get('nombre')
        medico.Apellido = request.POST.get('apellido')
        medico.especialidad = request.POST.get('especialidad')
        medico.Telefono = request.POST.get('telefono')
        medico.Email = request.POST.get('email')
        medico.Contraseña = request.POST.get('contrasena')
        medico.save()
        messages.success(request, 'Medico editado.')
        return redirect('medicos_admin')
    return render(request, 'Editar_medico.html', context={'medico':medico})

def editar_cita(request,id_cita):
    cita = Cita.objects.get(id_cita=id_cita)
    if request.method == 'POST':
        cita.medico = Medico.objects.get(Nombre=request.POST.get('Medico'))
        cita.paciente = Paciente.objects.get(Dni=request.POST.get('dni'))
        cita.fecha_cita = request.POST.get('Fechadedisponibilidad')
        cita.hora = request.POST.get('hora')
        cita.especialidad = request.POST.get('especialidad')
        cita.motivo = request.POST.get('Motivo')
        cita.estado = request.POST.get('estado')
        cita.save()
        messages.success(request, 'Cita editada.')
        return redirect('citas_admin')
    Medicos=Medico.objects.all()
    return render(request, 'Editar_cita.html', context={'cita':cita, 'medicos':Medicos})


def pdfAdmin_citas(request):
    template = get_template('citasForm.html')
    citas = Cita.objects.all()  # Ajusta esto según tu modelo y query
    context = {'citas': citas}
    html = template.render(context)
    
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return HttpResponse('Error al generar el PDF', status=400)

def excel_citas(request):
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Citas"

    # Definir los encabezados
    columns = ['#', 'Paciente', 'Especialista', 'Especialidad', 'Fecha', 'Estado']

    # Escribir los encabezados
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Obtener todas las citas
    citas = Cita.objects.all()

    # Escribir los datos
    for row_num, cita in enumerate(citas, 2):
        ws.cell(row=row_num, column=1, value=cita.id_cita)
        ws.cell(row=row_num, column=2, value=f"{cita.paciente.Nombre} {cita.paciente.Apellido}")
        ws.cell(row=row_num, column=3, value=f"{cita.medico.Nombre} {cita.medico.Apellido}")
        ws.cell(row=row_num, column=4, value=cita.medico.especialidad)
        ws.cell(row=row_num, column=5, value=cita.fecha_cita)
        ws.cell(row=row_num, column=6, value=cita.estado)

    # Ajustar el ancho de las columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Citas.xlsx'

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)

    return response

def pdfAdmin_pacientes(request):
    template = get_template('pacientesForm.html')
    pacientes = Paciente.objects.all()  # Ajusta esto según tu modelo y query
    context = {'pacientes': pacientes}
    html = template.render(context)
    
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return HttpResponse('Error al generar el PDF', status=400)

def excel_pacientes(request):
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Pacientes"

    # Definir los encabezados
    columns = ['#', 'DNI', 'Nombre', 'Apellido', 'Fecha de Nacimiento', 'Género', 'Dirección', 'Teléfono', 'Email']

    # Escribir los encabezados
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Obtener todos los pacientes
    pacientes = Paciente.objects.all()

    # Escribir los datos
    for row_num, paciente in enumerate(pacientes, 2):
        ws.cell(row=row_num, column=1, value=paciente.id_paciente)
        ws.cell(row=row_num, column=2, value=paciente.Dni)
        ws.cell(row=row_num, column=3, value=paciente.Nombre)
        ws.cell(row=row_num, column=4, value=paciente.Apellido)
        ws.cell(row=row_num, column=5, value=paciente.Fecha_nacimiento)
        ws.cell(row=row_num, column=6, value=paciente.Genero)
        ws.cell(row=row_num, column=7, value=paciente.Direccion)
        ws.cell(row=row_num, column=8, value=paciente.Telefono)
        ws.cell(row=row_num, column=9, value=paciente.Email)

    # Ajustar el ancho de las columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Pacientes.xlsx'

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)

    return response

def pdfAdmin_medicos(request):
    template = get_template('medicosForm.html')
    medicos = Medico.objects.all()  # Ajusta esto según tu modelo y query
    context = {'medicos': medicos}
    html = template.render(context)
    
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return HttpResponse('Error al generar el PDF', status=400)

def excel_medicos(request):
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Médicos"

    # Definir los encabezados
    columns = ['#', 'Nombre', 'Apellido', 'Especialidad', 'Teléfono', 'Email']

    # Escribir los encabezados
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Obtener todos los médicos
    medicos = Medico.objects.all()

    # Escribir los datos
    for row_num, medico in enumerate(medicos, 2):
        ws.cell(row=row_num, column=1, value=medico.id_medico)
        ws.cell(row=row_num, column=2, value=medico.Nombre)
        ws.cell(row=row_num, column=3, value=medico.Apellido)
        ws.cell(row=row_num, column=4, value=medico.especialidad)
        ws.cell(row=row_num, column=5, value=medico.Telefono)
        ws.cell(row=row_num, column=6, value=medico.Email)

    # Ajustar el ancho de las columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Médicos.xlsx'

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)

    return response

def enviar_email_html(asunto, template, contexto, destinatario):
    html_message = render_to_string(template, contexto)
    plain_message = strip_tags(html_message)
    email = EmailMultiAlternatives(
        subject=asunto,
        body=plain_message,
        from_email=None,
        to=[destinatario],
    )
    email.attach_alternative(html_message, "text/html")
    email.send()